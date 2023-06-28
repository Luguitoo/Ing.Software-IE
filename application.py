from flask import Flask, url_for, redirect, render_template, send_file, request, session
from config import DevConfig
from database.conexion import *
from database.models import *
from database import models

import sqlite3
import os
from datetime import datetime
from sqlalchemy import text
import sqlalchemy as db
from database.conexion import *
from database.models import *

#Excel
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pandas as pd
#Json
from flask import jsonify
import json
#database con sqlalchemy
from database import models
from sqlalchemy import text, update
from database.conexion import engine
from database.conexion import SessionLocal
from sqlalchemy import func
#formulas
from formulas import *

import asyncio
models.Base.metadata.create_all(bind=engine)
##variables globales
global cant_a
cant_a = 0
global cont
cont = 0
global alumnos
alumnos = []

application = app = Flask(__name__)

app.config.from_object(DevConfig)
##dbtest = sqlite3.connect('NombreDeLaDB.db')

#Variables que uso de forma temporal ya que despues se va a guardar en la db
alumnos=[] #para la cargar de alumnos (ver /loadSt)
data = []  #para la carga del historial de materias (ver /histAl)


def calcular_semestre(anio_ingreso, semestres_totales):
    fecha_actual = datetime.now()
    año_actual = fecha_actual.year
    mes_actual = fecha_actual.month
    semestres_transcurridos = (año_actual - anio_ingreso) * 2  # Suponiendo 2 semestres por año

    # Ajustar el semestre si el mes actual está en el segundo semestre
    if mes_actual > 6:
        semestres_transcurridos += 1
    if semestres_transcurridos > semestres_totales:
        return semestres_totales
    else:
        return semestres_transcurridos

import re
def extraer_numero(string):
    numero = re.match(r'^\d+', string)
    if numero:
        return int(numero.group())
    else:
        return 0

#leer el/los excel de las notas, guarda en la db y devuelve un js con los datos leidos
#si se carga varios excel solo devuelve el json del ultimo
def leer_excel_notas(archivos):
    session = SessionLocal()
    try:
        for archivo in archivos:
            df = pd.read_excel(archivo)
            filename = archivo.filename
            # Guardar el archivo XLSX
            try:
                os.mkdir('./static/temp')
            except FileExistsError:
                pass
            archivo_xlsx = f"./static/temp/{archivo.filename}.xlsx"
            df.to_excel(archivo_xlsx, index=False)
            wb = load_workbook('./static/temp/{a}.xlsx'.format(a = archivo.filename))
            ws = wb["Sheet1"]
            nomb = ws['C2'].value
            matr = ws['C3'].value
            matr = matr.split("/")[0].strip() # esto devuelve la segunda matrícula
            print(f"Alumno: {nomb}, matricula: {matr}")
            #elimina de ls bd todos las calificaciones anteriores
            #capaz update sea mejor
            historial = session.query(models.Historial).filter(models.Historial.matricula == matr).all() #Se corrigió un error en el que no se eliminaba el historial viejo
            if historial:
                    for i in historial:
                        session.delete(i)
                        session.flush()
            inicio = 4
            cont = 0
            
            b = True
            data = []  # Lista para almacenar los datos
            while b:
                if not ws['A{a}'.format(a=str(inicio + 1))].value:
                    cont += 1
                    if cont > 1:
                        b = False
                else:
                    mat = ws['A{a}'.format(a=str(inicio + 1))].value
                    cod = ws['D{a}'.format(a=str(inicio + 1))].value
                    opo = ws['E{a}'.format(a=str(inicio + 1))].value
                    nota = ws['F{a}'.format(a=str(inicio + 1))].value
                    nota = nota.split(':')[0].strip()
                    nota = extraer_numero(str(nota))
                    act = ws['G{a}'.format(a=str(inicio + 1))].value
                    fec = ws['H{a}'.format(a=str(inicio + 1))].value
                    data.append(
                    {
                        'alu': ws['C2'].value,
                        'matr': matr,
                        'mat': mat,
                        'cod': cod,
                        'opo': opo,
                        'nota': nota,
                        'act': act,
                        'fec': fec
                    })
                    #carga la nueva calificación a bd
                    date = datetime.strptime(fec, '%d/%m/%Y')
                    newhistorial = models.Historial(matricula = matr, materia_codigo = cod, nota = nota, oportunidad = opo, fecha_examen = date)
                    alumno = session.query(models.Alumnos).filter(models.Alumnos.matricula == matr).first()
                    now = datetime.now()
                    alumno.ult_act = now
                    session.add(newhistorial)
                    session.flush()
                inicio += 1
            #verifica su regularidad
            if alumno.estado_id != 3:
                cant_semestres = session.query(models.Semestre).count()
                cohorte = session.query(models.Cohortes).filter(models.Cohortes.cohorte_id == alumno.cohorte_id).first()
                sem_cursados = calcular_semestre(cohorte.cohorte_inicio, cant_semestres)
                anios = int(sem_cursados / 2) + cohorte.cohorte_inicio  #para saber cuantos años ya pasó de su ingreso
                if sem_cursados % 2 == 0: #pares si son de segundo semetre del año
                    filter = datetime.strptime("30-06-"+str(anios), "%d-%m-%Y") # tiene tiempo hasta el 30 de junio de rendir la extraodinaria si es necesario
                else:
                    filter = datetime.strptime("30-10-"+str(anios), "%d-%m-%Y") # tiene tiempo hasta el 30 de octubre de rendir la extraodinaria si es necesario
                if sem_cursados >= cant_semestres:
                    estado_sem(sem_cursados, matr, filter, session)
                    aprovados = session.query(models.Historial).filter(models.Historial.nota > 1, models.Historial.matricula == matr).count()
                    cant_materias = session.query(models.Materias).count()
                    if aprovados == cant_materias:
                        alumno.estado_id = 5 #titulado
                    else:
                        alumno.estado_id = 2 # irregular
                    session.flush()
                else:
                    alumno.estado_id = estado_sem(sem_cursados, matr, filter, session)
                    session.flush()
            os.remove('./static/temp/{a}.xlsx'.format(a = archivo.filename))
            
        
    except Exception as e:
        print(e, 'EE')
        try:
            os.remove('./static/temp/{a}.xlsx'.format(a = filename)) #elimina el excel del sistema
            session.close()
            return None
        except:
            session.close()
            return None
    session.commit()
    session.close()
    return data

##funcion para recorrer semestres y asignar el estado del alumno en cada uno y devolver el último estado
def estado_sem(sem_cursados, matr, filter, session):
    estados_semestrales = session.query(models.Est_Sem_Alumnos).filter(models.Est_Sem_Alumnos.matricula == matr).all()
    mayor_semestre_rendido = session.query(func.max(models.Materias.semestre_id)).join(models.Historial, models.Historial.materia_codigo == models.Materias.materia_codigo).filter(models.Historial.matricula == matr).scalar()
    if estados_semestrales:
            for i in estados_semestrales:
                session.delete(i)
                session.flush()
    for i in range(1,sem_cursados+1):
        aprovados = session.query(models.Historial).join(models.Materias).filter(models.Historial.matricula == matr, models.Historial.fecha_examen < filter, models.Historial.nota > 1, models.Materias.semestre_id <= i).count() #cant de materias rendidas y aprovadas antes de la fecha limite
        cant_materias = session.query(models.Materias).filter(models.Materias.semestre_id <= i).count()
        if i <= mayor_semestre_rendido:
            if aprovados >= cant_materias:
                estado = 1 ##Regular
            else:
                estado = 2 ##Irregular
        else:
            estado = 6 ##Abandonado
        newestadosem = models.Est_Sem_Alumnos(matricula = matr, semestre_id = i, estado_id = estado)
        session.add(newestadosem)
        session.flush()
    return estado

##funcion agregar alumno, pq puede dar error
def agregar_alumno(matricula, nombre, id_cohor, convalidado, session):
    if matricula:
        #session = SessionLocal()
        # Realizar alguna advertencia si es que ya existe el alumno
        matricula = matricula.split("/")[0].strip()
        alumno = session.query(models.Alumnos).filter(models.Alumnos.matricula == matricula).first()
        if convalidado == 1:
            print(1)
            estado = 3 #estado convalidado
        elif alumno:
            print(2)
            estado = alumno.estado_id
        else:
            print(3)
            estado = 1 #alumno regular
        if alumno:
            print(4)
            alumno.alumno_nombre = nombre
            alumno.cohorte_id = id_cohor
            alumno.estado_id = estado
            session.commit()
        else:
            print(5)
            print(matricula)
            print(nombre)
            print(id_cohor)
            print(estado)
            new_alumno = models.Alumnos(matricula=matricula, alumno_nombre=nombre, cohorte_id=id_cohor, estado_id = estado)
            session.add(new_alumno)
            session.commit()
        #session.close
        return True

#verifica si existen materias cargadas en el sistema antes de acceder a caulquier url
@app.before_request
def verif_materias():
    session = SessionLocal()
    current_endpoint = request.endpoint
    models.insert_estados(session) #Crea los estados en la base de datos
    #models.insert_test_data(session)
    if current_endpoint != 'cargadematerias' and current_endpoint != 'read_materias' and current_endpoint != 'static':
        materias = session.query(models.Materias).all()
        if not materias:
            app.jinja_env.globals['alert_message'] = "Antes de usar el sistema debe de cargar el listado de asignaturas"
            session.close()
            return redirect('/cargar_materias')
    session.close()

@app.route('/')
def index():
    session = SessionLocal()
    cohortes = session.query(models.Cohortes).all()
    return render_template('index.html', data=cohortes)

@app.route('/cohortes')
def cohortes():
    session = SessionLocal()
    cohorte = session.query(models.Cohortes.cohorte_id).all()
    if cohorte:
        session.close()
        return redirect('/selCoh?cid='+str(cohorte[0][0]))
    else:
        session.close()
        return redirect('/')
    

@app.route('/selCoh', methods=['GET'])
def selCoh():
    session = SessionLocal()
    if request.method == 'GET':
        cid = request.args.get('cid')
        value = {'x': cid}

        con = text("select * from alumnos where alumnos.cohorte_id = :x")
        alumnos = session.execute(con, value)
        cohortes = session.query(models.Cohortes).all()
        cohorte = session.query(models.Cohortes).filter(models.Cohortes.cohorte_id==cid).first()
        session = SessionLocal()
        return render_template('cohortes.html', data=alumnos, coh=cohorte, cohortes = cohortes, coh_id = cid)

    #for i in range(1,8):
    #    print(tasa_promocion_semestral(1, i, session))

    """for i in outs:
        print(i.matricula, ' - ', i.nota)"""
    cohortes = session.query(models.Cohortes).all()
    session.close()
    return render_template('index.html', cohortes=cohortes)

@app.route('/notas')
def notas():
    global cant_a
    global cont
    global alumnos
    print(alumnos)
    if cont == 0:
        cont += 1
    else:
        cant_a -=1 
        cont += 1
    print(cant_a)
    print(cont)
    return render_template('notas.html', cant_a = cant_a, alumno = alumnos[cont - 1])

##Procesos
##Ruta de descarga del modelo del excel, terminado Lugo
@app.route('/download_template', methods = ['GET', 'POST'])
def download_template():
    if request.method == "POST":
        return send_file('./static/resources/IE-CyT.xlsx', as_attachment=True)

@app.route('/download_template_mat', methods = ['GET', 'POST'])
def download_template_mat():
        return send_file('./static/resources/materias.xlsx', as_attachment=True) 
    
##Leer excel de listado de alumnos y carga en el front, terminado Lugo
@app.route('/read_excel', methods=['POST']) 
def read_excel():
    global cant_a
    global cont
    global alumnos
    alumnos = []
    cant_a = 0
    cont = 0
    session = SessionLocal()
    archivo = request.files['archivo']
    inic = request.form['desde']
    fin = request.form['hasta']
    if "archivo" not in request.files:
        print("No se envió ningún archivo")
        return "No se envió ningún archivo"
    elif archivo.filename == "":
        print("No se seleccionó ningún archivo")
        return "No se seleccionó ningún archivo"
    
    #Cargamos el archivo
    wb = load_workbook(archivo)
    ws = wb["Hoja1"]
    inicio = 1
    b = True
    data = []  # Lista para almacenar los datos

    #Carga la cohorte si no existe
    id_cohor = session.query(models.Cohortes.cohorte_id).filter(models.Cohortes.cohorte_inicio == inic).first()
    if not id_cohor: #Falta alguna advertencia si es que ya existe la cohorte
        newcohorte = models.Cohortes(cohorte_inicio = inic,cohorte_fin = fin)
        session.add(newcohorte)
        session.flush()
        id_cohor = session.query(models.Cohortes.cohorte_id).filter(models.Cohortes.cohorte_inicio == inic).first()
    """else:
        #si ya existe la cohorte elimina todos los alumnos con esa cohorte para luego poder actualizar su listado
        coh = session.query(models.Alumnos).filter(models.Alumnos.cohorte_id == id_cohor[0]).all()
        for i in coh:
            session.delete(i)
            session.commit()"""
    matriculas = []
    print('N, Matr, Name')
    while b:
        if not ws['A{a}'.format(a=str(inicio + 1))].value:
            b = False
        else:
            print(ws['A{a}'.format(a = str(inicio + 1))].value, ws['B{a}'.format(a = str(inicio + 1))].value, ws['D{a}'.format(a = str(inicio + 1))].value)
            num = ws['A{a}'.format(a=str(inicio + 1))].value
            matricula = ws['B{a}'.format(a=str(inicio + 1))].value
            matricula = matricula.split("/")[0].strip()
            nombre = ws['D{a}'.format(a=str(inicio + 1))].value
            convalidado = ws['H{a}'.format(a=str(inicio + 1))].value
            alumnos.append(nombre)
            data.append({
                'num': num,
                'matricula': matricula,
                'nombre': nombre,
                'convalidado': convalidado
            })
            print(convalidado)
            matriculas.append(matricula)
            inicio += 1
            #matric = session.query(models.Alumnos.matricula).filter(models.Alumnos.matricula == matricula).scalar()
            # Llamar a la función agregar_alumno dentro de un bucle de eventos asyncio
            alu = agregar_alumno(matricula, nombre, id_cohor[0], convalidado, session=session)
            if alu == None:
                return "El Alumno con matricula "+ str(matricula) + " ya está registrando en otra cohorte", 400
    #elimina los alumnos que no vinieron en el excel junto con su historial
    alumnos = session.query(models.Alumnos).filter(models.Alumnos.cohorte_id == id_cohor[0]).all()
    for alumno in alumnos:
        if not alumno.matricula in matriculas:
            historial = session.query(models.Historial).filter(models.Historial.matricula == alumno.matricula).all()
            if historial:
                for i in historial:
                    session.delete(i)
                    session.flush()
            session.delete(alumno)
            session.flush()
    cant_a = inicio - 1
    cant_matriculados = session.query(models.Cantidad_inscript).filter(models.Cantidad_inscript.cohorte_id == id_cohor[0]).first()
    if cant_matriculados:
        cant_matriculados.cantidad = cant_a
        session.flush()
    else:
        cant_matriculados = models.Cantidad_inscript(cohorte_id = id_cohor[0], semestre_id = 1, cantidad = cant_a)
        session.add(cant_matriculados)
    ##print(cant_a)
    session.commit()
    session.close()

    # Convertir a JSON
    json_data = jsonify(data)
    return json_data

##Leer notas del excel, terminado lugo
@app.route('/read_all_notas', methods=['POST'])
def read_all_notas():
    session = SessionLocal()
    archivos = request.files.getlist("archivo")
    if "archivo" not in request.files:
        print("No se envió ningún archivo")
        return "No se envió ningún archivo"
    
    #archivo = request.files['archivo']
    result = leer_excel_notas(archivos)
    if result == None:
        print('Ocurrio un error al procesar el documentos')
        return 'Ocurrio un error al procesar el documentos', 500
    id_cohorte = session.query(models.Alumnos.cohorte_id).filter(models.Alumnos.matricula==result[0]['matr']).first()
    session.commit()
    session.close()
    return redirect('/selCoh?cid='+str(id_cohorte[0]))

@app.route('/read_notas', methods=['POST'])
def read_notas():
    session = SessionLocal()
    if "archivo" not in request.files:
        print("No se envió ningún archivo")
        return "No se envió ningún archivo", 400
    archivo = request.files.getlist("archivo")
    if archivo[0].filename == "":
        print("No se seleccionó ningún archivo")
        return "No se seleccionó ningún archivo", 400
    print(archivo[0].filename)
    #si devuelve none es porque ocurrio un error en la lectura
    data = leer_excel_notas(archivo)
    if data == None:
        print("Ocurrió un error al leer el documento")
        return "Ocurrió un error al leer el documento", 500
    #coomit de la bd y cierre de sesión
    json_data = jsonify(data)
    session.commit()
    session.close()
    return json_data

#Ver historial de asignaturas del alumno
@app.route('/historial/<mat>')
def historial(mat:str):
    session = SessionLocal()
    semestres =  session.query(func.count(models.Semestre.semestre_id)).first()
    materias = session.query(models.Materias.materia_codigo, models.Materias.materia_descrip, models.Materias.semestre_id).order_by(models.Materias.semestre_id).all()
    historial = []
    for materia in materias:
        calif = session.query(models.Historial).filter(models.Historial.matricula==mat.upper(), models.Historial.materia_codigo == materia[0]).all()
        estado = {
            'codigo': materia[0],
            'descripcion': materia[1],
            'semestre': materia[2],
        }
        print(calif)
        if(calif == []):
            estado['estado'] = 'Sin registros'
        elif len(calif) >= 1:
            may =extraer_numero(str(calif[0].nota))
            for i in range(len(calif)):
                nota =  extraer_numero(str(calif[i].nota))
                if nota > may:
                    may = nota
            if may > 1:
                estado['estado'] = 'Aprobado'
            else:
                estado['estado'] = 'No aprobado'
        historial.append(estado)
    alumno = session.query(models.Alumnos).filter(models.Alumnos.matricula == mat).first()
    session.close()
    print(historial)
    return render_template('historial.html', historial = historial, alumno = alumno)


@app.route('/cargar_materias')
def cargadematerias():
    session = SessionLocal()
    materias = session.query(models.Materias).order_by(models.Materias.semestre_id).all()
    session.close()
    #mensaje de error en pantalla
    alert_message = app.jinja_env.globals.get('alert_message')
    app.jinja_env.globals.pop('alert_message', None)
    return render_template('cargarmaterias.html', materias=materias, alert_message = alert_message)

@app.route('/read_materias', methods=['POST'])
def read_materias(): 
    if request.method == 'POST':
        cant_a = 0
        cont = 0
        session = SessionLocal()
        archivo = request.files['archivo']
        if "archivo" not in request.files:
            print("No se envió ningún archivo")
            return "No se envió ningún archivo"
        elif archivo.filename == "":
            print("No se seleccionó ningún archivo")
            return "No se seleccionó ningún archivo"
        #Cargamos el archivo
        wb = load_workbook(archivo)
        ws = wb["Hoja1"]
        inicio = 1
        list = []
        b = True
        print("Todo bien")

        print('Cod, Mate, Semes')
        semestres = session.query(models.Semestre).all()
        for i in semestres:
            session.delete(i)
            session.flush()
        materias = session.query(models.Materias).all()
        for i in materias:
            session.delete(i)
            session.flush()
        may = 0
        while b:
            if not ws['A{a}'.format(a=str(inicio + 1))].value:
                b = False
            else:
                print(ws['A{a}'.format(a = str(inicio + 1))].value, ws['B{a}'.format(a = str(inicio + 1))].value, ws['C{a}'.format(a = str(inicio + 1))].value)
                codm = ws['A{a}'.format(a=str(inicio + 1))].value
                nom = ws['B{a}'.format(a=str(inicio + 1))].value
                semes = ws['C{a}'.format(a=str(inicio + 1))].value
                list.append({
                    'cod': codm,
                    'materia': nom,
                    'semestre': semes,
                })
                inicio += 1
                if semes < 1:
                    return "Error, los semestres no pueden ser menores a 1", 400
                if semes > may:
                    may = semes
                cargmateria = models.Materias(materia_codigo = codm, materia_descrip = nom, semestre_id = semes)
                session.add(cargmateria)
                session.flush()
        cant_a = inicio - 1
        semes = session.query(models.Semestre).all()
        for i in semes:
            session.delete(semes)
            session.flush()
        for i in range(may):
            cant_semes = models.Semestre(semestre_id = i+1)
            session.add(cant_semes)
        session.commit()
        session.close()
        json_data = jsonify(list)
        return json_data

@app.route("/salidas", methods=['GET', 'POST'])
def salidas():
    session = SessionLocal()
    cohortes = session.query(models.Cohortes.cohorte_id,models.Cohortes.cohorte_inicio,models.Cohortes.cohorte_fin).all()
    semestres = session.query(models.Semestre.semestre_id).all()
    session.close()
    if request.method == "POST":
        cohorte_id = request.form['cohorte_id']
        semestre_inicio = int(request.form['semestre_inicio'])
        semestre_fin = int(request.form['semestre_fin'])
        #print(cohorte_id,semestre_inicio,semestre_fin)
        respuesta = []
        anual = []
        semestral = []
        for i in range(semestre_inicio//2 + 1, semestre_fin//2 + 1):
          anual.append({
              "anho" : i,
              i: tasa_promocion_anual(cohorte_id,i,session),
          })

        for i in range(semestre_inicio, semestre_fin + 1):
          semestral.append({
              "semestre" : i,
              i: {
                  "desercion" : tasa_desercion_semestral(cohorte_id,i,session),
                  "retencion" : tasa_retencion(cohorte_id,i,session),
              },
          })
        respuesta.append({
            "eficiencia" : eficiencias(cohorte_id,session),
            "desercion_generacional" : tasa_desercion_generacional(cohorte_id, session),
            "promocion_semestral" : tasa_promocion_semestral(cohorte_id, semestre_inicio, semestre_fin, session),
            "anuales" : anual,
            "semestrales" : semestral,
        })


        json_data = json.loads(json.dumps(respuesta))
        #json_data = jsonify(respuesta)
        #print(json_data)

        # Mostrar la cadena json por pantalla
        return render_template('salidas.html', cohortes = cohortes, semestres = semestres, 
                               json_data = respuesta[0], cohorte_id = cohorte_id, 
                               semestre_inicio = semestre_inicio, semestre_fin = semestre_fin)
    else:
        return render_template('salidas.html', cohortes = cohortes, semestres = semestres, 
                               json_data = "", sel_cohorte = "", 
                               semestre_inicio = "", semestre_fin = "")

@app.route('/cant_inscriptos/<int:id>')
def cant_inscriptos(id):
    session = SessionLocal()
    datos = []
    semestre = session.query(models.Semestre).count()
    id_cohorte = session.query(models.Cantidad_inscript.cohorte_id).filter(models.Cantidad_inscript.cohorte_id == id, models.Cantidad_inscript.semestre_id == 1).scalar()
    #Verifica si realmente existe esa cohorte en la bd
    cohorte = session.query(models.Cohortes.cohorte_id).filter(models.Cohortes.cohorte_id == id).scalar()
    if cohorte:
        if id_cohorte:
            #si ya existe el primero, inicializa el resto en cero
            for x in range(2, semestre + 1):
                id_semestre = session.query(models.Cantidad_inscript).filter(models.Cantidad_inscript.cohorte_id == id, models.Cantidad_inscript.semestre_id == x).scalar()
                #solo se inicializa si aún no existen
                if not id_semestre:
                    nuevo = models.Cantidad_inscript(cohorte_id = id, semestre_id = x, cantidad = 0 )          
                    session.add(nuevo)
                    session.commit()
        else:
            print("aún no se creo el primer registro")
    datos = session.query(models.Cantidad_inscript.semestre_id, models.Cantidad_inscript.cantidad).filter(models.Cantidad_inscript.cohorte_id == id).all()
    print(datos) 
    session.close()
    return render_template("cant_incriptos.html", datos=datos, id=id)
@app.route("/actualizar_cantidad", methods=['POST'])
def actualizar_cantidad():
    session = SessionLocal()
    semestre = session.query(models.Semestre).count()
    id = request.form["id_cohorte"]
    #Verifica si realmente existe esa cohorte en la bd
    cohorte = session.query(models.Cohortes.cohorte_id).filter(models.Cohortes.cohorte_id == id).scalar()
    if cohorte:
        for x in range(1, semestre + 1):
            try:
                cant = request.form["cant_"+str(x)]
                datos = session.query(Cantidad_inscript).filter_by(cohorte_id=id, semestre_id=x).scalar()  
                datos.cantidad = cant
                session.commit()
            #para controlar el error cuando no hay campos para completar
            except:
                return("Error")
    else:
        print("no existe esta cohorte")
    session.close()
    return redirect("/")

if __name__=='__main__':
    app.run(debug = True, port= 8000)

